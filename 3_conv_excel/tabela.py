'''modulo da openpyxl'''
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

#carrega o dataframe para um workbook
def gerar_tabela(filename):
    wb = load_workbook(filename)
    ws = wb.active
    ws.title = 'Planilha de Consumidores'

    #ajusta a largura das colunas para que fiquem rente ao conteúdo
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    #intervalo de células
    min_col = ws.min_column
    max_col = ws.max_column
    min_row = ws.min_row
    max_row = ws.max_row
    cell_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"

    #cria uma tabela
    table = Table(displayName="tabela_produtos", ref=cell_range)

    #estilo da tabela
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    #adiciona a tabela a planilha
    ws.add_table(table)

    wb.save(filename)

    return table

